"""Local document text extraction."""

from __future__ import annotations

import csv
import io
import json
from zipfile import BadZipFile

from app.providers.base import DocIntelligenceProvider


class LocalDocIntelligenceProvider(DocIntelligenceProvider):
    def extract_text(self, data: bytes, *, filename: str, content_type: str = "") -> str:
        name = filename.lower()
        if name.endswith(".txt") or content_type.startswith("text/plain"):
            return data.decode("utf-8", errors="replace")
        if name.endswith(".json") or "json" in content_type:
            try:
                parsed = json.loads(data.decode("utf-8"))
                return json.dumps(parsed, indent=2)
            except json.JSONDecodeError:
                return data.decode("utf-8", errors="replace")
        if name.endswith(".csv"):
            text = data.decode("utf-8", errors="replace")
            reader = csv.reader(io.StringIO(text))
            return "\n".join([", ".join(row) for row in reader])
        if name.endswith(".docx"):
            try:
                from docx import Document

                doc = Document(io.BytesIO(data))
                return "\n".join(p.text for p in doc.paragraphs if p.text.strip())
            except (BadZipFile, Exception):
                return data.decode("utf-8", errors="replace")
        if name.endswith((".xlsx", ".xls")):
            try:
                from openpyxl import load_workbook

                wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
                lines: list[str] = []
                for ws in wb.worksheets:
                    for row in ws.iter_rows(values_only=True):
                        lines.append(", ".join("" if c is None else str(c) for c in row))
                return "\n".join(lines)
            except Exception:
                return data.decode("utf-8", errors="replace")
        return data.decode("utf-8", errors="replace")
